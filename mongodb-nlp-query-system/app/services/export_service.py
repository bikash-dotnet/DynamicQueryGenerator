"""
Export service for CSV and Excel file generation
"""

import csv
import io
import json
import zipfile
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

from app.config import settings
from app.database.data_repository import data_repository, convert_objectid_to_str
from app.database.query_repository import query_repository

logger = logging.getLogger(__name__)


class ExportService:
    """Handle data export to CSV and Excel formats"""
    
    def __init__(self):
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)
    
    async def export_query_to_csv(self, query_id: str, query_text: Optional[str] = None) -> Optional[str]:
        """
        Export query results to CSV format
        
        Args:
            query_id: ID of the stored query
            query_text: Original query text (for filename)
        
        Returns:
            Path to generated file or None if failed
        """
        try:
            # Retrieve the stored query
            stored_query = await query_repository.find_by_hash(query_id)
            if not stored_query:
                logger.error(f"Query not found: {query_id}")
                return None
            
            # Execute the query to get all results (no limit)
            results, total_count = await data_repository.execute_query(
                stored_query.collection_name,
                stored_query.generated_query,
                limit=settings.MAX_RESULTS  # Use max results limit
            )
            
            if not results:
                logger.warning(f"No results to export for query: {query_id}")
                return None
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}_{query_id[:8]}.csv"
            
            # Generate CSV
            file_path = await self._generate_csv(results, filename)
            
            logger.info(f"CSV export completed: {filename} with {len(results)} records")
            return file_path
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return None
    
    async def export_query_to_excel(self, query_id: str, query_text: Optional[str] = None) -> Optional[str]:
        """
        Export query results to Excel format
        
        Args:
            query_id: ID of the stored query
            query_text: Original query text (for filename)
        
        Returns:
            Path to generated file or None if failed
        """
        try:
            # Retrieve the stored query
            stored_query = await query_repository.find_by_hash(query_id)
            if not stored_query:
                logger.error(f"Query not found: {query_id}")
                return None
            
            # Execute the query to get all results
            results, total_count = await data_repository.execute_query(
                stored_query.collection_name,
                stored_query.generated_query,
                limit=settings.MAX_RESULTS
            )
            
            if not results:
                logger.warning(f"No results to export for query: {query_id}")
                return None
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}_{query_id[:8]}.xlsx"
            
            # Generate Excel
            file_path = await self._generate_excel(results, filename)
            
            logger.info(f"Excel export completed: {filename} with {len(results)} records")
            return file_path
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return None
    
    async def _generate_csv(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Generate CSV file from data
        
        Args:
            data: List of dictionaries to export
            filename: Output filename
        
        Returns:
            Path to generated file
        """
        file_path = self.export_dir / filename
        
        if not data:
            # Create empty CSV with headers only
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([])
            return str(file_path)
        
        # Flatten nested data
        flattened_data = [self._flatten_json(row) for row in data]
        
        # Get all unique columns
        columns = set()
        for row in flattened_data:
            columns.update(row.keys())
        columns = sorted(columns)
        
        # Write CSV with proper encoding for Excel compatibility
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=columns, quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            
            for row in flattened_data:
                # Ensure all columns exist in row
                clean_row = {col: row.get(col, '') for col in columns}
                writer.writerow(clean_row)
        
        # Check file size and compress if needed
        file_size = file_path.stat().st_size / (1024 * 1024)
        if file_size > 10:  # Compress if > 10MB
            zip_path = await self._compress_file(file_path)
            return zip_path
        
        return str(file_path)
    
    async def _generate_excel(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Generate formatted Excel file from data
        
        Args:
            data: List of dictionaries to export
            filename: Output filename
        
        Returns:
            Path to generated file
        """
        file_path = self.export_dir / filename
        
        if not data:
            # Create empty workbook
            wb = Workbook()
            # Ensure we have an active sheet, fallback to create_sheet if active is None
            ws = wb.active if wb.active is not None else wb.create_sheet()
            ws.title = "Query Results"
            wb.save(file_path)
            return str(file_path)
        
        # Flatten data for Excel
        flattened_data = [self._flatten_json(row) for row in data]
        df = pd.DataFrame(flattened_data)
        
        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Handle large datasets - split into multiple sheets
            rows_per_sheet = 100000
            if len(df) > rows_per_sheet:
                num_sheets = (len(df) + rows_per_sheet - 1) // rows_per_sheet
                for i in range(num_sheets):
                    start = i * rows_per_sheet
                    end = min((i + 1) * rows_per_sheet, len(df))
                    sheet_name = f"Part_{i+1}_{start+1}_{end}"
                    df.iloc[start:end].to_excel(writer, sheet_name=sheet_name[:31], index=False)
            else:
                df.to_excel(writer, sheet_name="Query Results", index=False)
            
            # Apply formatting to all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                self._format_excel_worksheet(worksheet)
        
        # Compress if too large
        file_size = file_path.stat().st_size / (1024 * 1024)
        if file_size > 10:
            zip_path = await self._compress_file(file_path)
            return zip_path
        
        return str(file_path)
    
    def _format_excel_worksheet(self, worksheet):
        """Apply professional formatting to Excel worksheet"""
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Auto-size columns
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format data rows
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Format dates if detected
                if isinstance(cell.value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    async def _compress_file(self, file_path: Path) -> str:
        """Compress file to ZIP archive"""
        zip_path = file_path.with_suffix('.zip')
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(file_path, file_path.name)
        
        # Remove original file
        file_path.unlink()
        
        logger.info(f"Compressed to: {zip_path.name}")
        return str(zip_path)
    
    def _flatten_json(self, data: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
        """
        Flatten nested JSON objects for CSV/Excel export
        
        Example: {"user": {"name": "John"}} -> {"user_name": "John"}
        """
        flattened = {}
        
        for key, value in data.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key
            
            if isinstance(value, dict):
                flattened.update(self._flatten_json(value, new_key, sep=sep))
            elif isinstance(value, list):
                # Convert list to JSON string for better readability
                if value and all(isinstance(item, (str, int, float)) for item in value):
                    flattened[new_key] = ', '.join(str(item) for item in value)
                else:
                    flattened[new_key] = json.dumps(value)
            elif isinstance(value, datetime):
                flattened[new_key] = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, (int, float, bool)):
                flattened[new_key] = value
            else:
                flattened[new_key] = str(value) if value else ''
        
        return flattened
    
    async def get_file(self, filename: str) -> Optional[Path]:
        """Get file path by filename"""
        file_path = self.export_dir / filename
        if file_path.exists():
            return file_path
        return None
    
    async def delete_file(self, filename: str) -> bool:
        """Delete an export file"""
        file_path = self.export_dir / filename
        if file_path.exists():
            file_path.unlink()
            logger.info(f"Deleted file: {filename}")
            return True
        return False
    
    async def cleanup_old_files(self, max_age_hours: int = 24) -> int:
        """Delete files older than max_age_hours"""
        import time
        deleted_count = 0
        current_time = time.time()
        
        for file_path in self.export_dir.iterdir():
            if file_path.is_file():
                file_age_hours = (current_time - file_path.stat().st_mtime) / 3600
                if file_age_hours > max_age_hours:
                    file_path.unlink()
                    deleted_count += 1
        
        if deleted_count:
            logger.info(f"Cleaned up {deleted_count} old files")
        
        return deleted_count


# Singleton instance
export_service = ExportService()